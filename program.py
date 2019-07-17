import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell
from os import listdir
from os.path import isfile, join
from tkinter import filedialog, Tk

def get_files():
    mypath = "C:/Users/EX002236/Desktop/openpyxl"
    files = [f for f in listdir(mypath) if isfile(join(mypath, f))]
    return files

class Color:
    def __init__(self):
        self.key = True
        while self.key:
            root = Tk()
            root.withdraw()
            root.filename =  filedialog.askopenfilename(initialdir = "/",title = "Select file",filetypes = (("jpeg files","*.xlsx"),("all files","*.*")))
            if root.filename == "":
                self.key = True
            else:
                break

        self.name = root.filename
        self.book = openpyxl.load_workbook(self.name)
        self.sheet = self.book.active
        self.redFill = PatternFill(start_color="ffab94",
                              end_color="ffab94",
                              fill_type='solid')
        self.greenFill = PatternFill(start_color="c1dac2",
                              end_color="c1dac2",
                              fill_type='solid')

    def main(self):
        print("running...")
        for row in range(len(self.sheet["X"])):
            cell = self.sheet.cell(row=row + 1, column=24)
            if cell.value == "open":
                for column in range(24):
                    self.sheet.cell(row=row + 1, column=column + 1).fill = self.redFill
            elif cell.value == "closed":
                for column in range(24):
                    self.sheet.cell(row=row + 1, column=column + 1).fill = self.greenFill

        self.book.save(self.name)
        print("done")




color = Color()
color.main()
